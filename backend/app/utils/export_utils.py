# -*- coding: utf-8 -*-
"""
导出工具类
提供PDF和Excel导出功能
"""

import io
import json
from datetime import datetime
from typing import List, Dict, Any

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportUtils:
    """导出工具类"""
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str = None) -> io.BytesIO:
        """导出数据到Excel格式
        
        Args:
            data: 要导出的数据列表
            filename: 文件名（可选）
            
        Returns:
            BytesIO: Excel文件的字节流
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl库未安装，无法导出Excel格式")
            
        if not data:
            raise ValueError("导出数据不能为空")
            
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "报表数据"
        
        # 设置标题样式
        title_font = Font(bold=True, size=12)
        title_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 获取列名
        if isinstance(data[0], dict):
            columns = list(data[0].keys())
        else:
            columns = [f"列{i+1}" for i in range(len(data[0]))]
            
        # 写入标题行
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
            
        # 写入数据行
        for row_idx, row_data in enumerate(data, 2):
            if isinstance(row_data, dict):
                for col_idx, column in enumerate(columns, 1):
                    value = row_data.get(column, "")
                    # 处理特殊数据类型
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value, ensure_ascii=False)
                    elif isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                for col_idx, value in enumerate(row_data, 1):
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value, ensure_ascii=False)
                    elif isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # 保存到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_to_pdf(data: List[Dict[str, Any]], title: str = "报表数据", filename: str = None) -> io.BytesIO:
        """导出数据到PDF格式
        
        Args:
            data: 要导出的数据列表
            title: PDF标题
            filename: 文件名（可选）
            
        Returns:
            BytesIO: PDF文件的字节流
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab库未安装，无法导出PDF格式")
            
        if not data:
            raise ValueError("导出数据不能为空")
            
        # 创建PDF文档
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        
        # 获取样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # 居中对齐
        )
        
        # 构建PDF内容
        story = []
        
        # 添加标题
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        # 准备表格数据
        if isinstance(data[0], dict):
            columns = list(data[0].keys())
            table_data = [columns]  # 标题行
            
            for row_data in data:
                row = []
                for column in columns:
                    value = row_data.get(column, "")
                    # 处理特殊数据类型
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value, ensure_ascii=False)
                    elif isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    row.append(str(value))
                table_data.append(row)
        else:
            # 如果数据不是字典格式，直接使用
            table_data = [[f"列{i+1}" for i in range(len(data[0]))]]
            for row_data in data:
                row = []
                for value in row_data:
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value, ensure_ascii=False)
                    elif isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    row.append(str(value))
                table_data.append(row)
        
        # 创建表格
        table = Table(table_data)
        
        # 设置表格样式
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # 添加生成时间
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        
        # 构建PDF
        doc.build(story)
        output.seek(0)
        return output
    
    @staticmethod
    def get_export_filename(format_type: str, base_name: str = "report") -> str:
        """生成导出文件名
        
        Args:
            format_type: 文件格式类型 ('excel' 或 'pdf')
            base_name: 基础文件名
            
        Returns:
            str: 完整的文件名
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = "xlsx" if format_type.lower() == "excel" else "pdf"
        return f"{base_name}_{timestamp}.{extension}"