from flask import request, jsonify, make_response
from app.api.v2 import v2_bp
from app.utils.auth_decorators import token_required
from app.utils.tenant_context import TenantContext, tenant_required
from app import db
from app.models.result import Result
from app.models.task import Task
from sqlalchemy import func
from datetime import datetime, timedelta
import json
import traceback
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def _calc_time_range(time_range: str):
    end_time = datetime.now()
    if time_range == '1h':
        start_time = end_time - timedelta(hours=1)
        interval = timedelta(minutes=5)
    elif time_range == '1d':
        start_time = end_time - timedelta(days=1)
        interval = timedelta(hours=1)
    elif time_range == '7d':
        start_time = end_time - timedelta(days=7)
        interval = timedelta(days=1)
    elif time_range == '30d':
        start_time = end_time - timedelta(days=30)
        interval = timedelta(days=1)
    else:
        start_time = end_time - timedelta(days=1)
        interval = timedelta(hours=1)
    return start_time, end_time, interval


def _format_time_label(dt: datetime, time_range: str) -> str:
    if time_range in ['1h', '1d']:
        return dt.strftime('%H:%M')
    return dt.strftime('%m-%d')


def _calculate_performance_stats(tenant_id: str, start_time: datetime, end_time: datetime):
    """计算性能指标统计"""
    try:
        # 计算时间间隔
        time_range = request.args.get('time_range', '1d', type=str)
        _, _, interval = _calc_time_range(time_range)
        
        # 查询所有结果数据
        results = db.session.query(Result, Task.type).join(
            Task, Result.task_id == Task.id
        ).filter(
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
            # 移除status过滤，包含所有结果来统计丢包率
        ).all()
        
        # 初始化统计数据
        tcp_times = []
        ping_times = []
        http_times = []
        api_times = []
        packet_losses = []
        
        # 按时间分组的丢包率数据
        time_grouped_packet_loss = {}
        
        # 解析每个结果的details字段
        for result, task_type in results:
            if not result.details:
                continue
                
            try:
                details = json.loads(result.details) if isinstance(result.details, str) else result.details
                
                # 根据任务类型提取相应的性能指标
                if task_type == 'tcp':
                    # TCP连接时间 - 只统计成功的结果
                    if result.status == 'success':
                        connect_time = details.get('connect_time') or details.get('response_time') or result.response_time
                        if connect_time and connect_time > 0:
                            tcp_times.append(float(connect_time))
                        
                elif task_type == 'ping':
                    # Ping延迟时间 - 只统计成功的结果
                    if result.status == 'success':
                        rtt_avg = details.get('rtt_avg') or details.get('execution_time') or result.response_time
                        if rtt_avg and rtt_avg > 0:
                            ping_times.append(float(rtt_avg))
                    
                    # 丢包率 - 统计所有结果（包括失败的）
                    packet_loss = details.get('packet_loss')
                    if packet_loss is not None:
                        packet_losses.append(float(packet_loss))
                        # 按时间分组 - 将结果时间对齐到时间间隔
                        result_time = result.created_at
                        # 计算该结果应该归属的时间段
                        time_diff = result_time - start_time
                        interval_count = int(time_diff.total_seconds() // interval.total_seconds())
                        aligned_time = start_time + interval * interval_count
                        time_key = _format_time_label(aligned_time, time_range)
                        
                        if time_key not in time_grouped_packet_loss:
                            time_grouped_packet_loss[time_key] = []
                        time_grouped_packet_loss[time_key].append(float(packet_loss))
                        
                elif task_type == 'http':
                    # HTTP响应时间 - 只统计成功的结果
                    if result.status == 'success':
                        response_time = details.get('response_time') or result.response_time
                        if response_time and response_time > 0:
                            http_times.append(float(response_time))
                        
                elif task_type == 'api':
                    # API响应时间 - 只统计成功的结果
                    if result.status == 'success':
                        total_time = details.get('total_time') or details.get('response_time') or result.response_time
                        if total_time and total_time > 0:
                            api_times.append(float(total_time))
                        
            except (json.JSONDecodeError, ValueError, TypeError):
                continue
        
        # 计算统计指标
        def calc_stats(data_list):
            if not data_list:
                return [0, 0, 0]  # [min, avg, max]
            return [
                round(min(data_list), 2),
                round(sum(data_list) / len(data_list), 2),
                round(max(data_list), 2)
            ]
        
        # 生成时间序列的丢包率数据
        packet_loss_time_series = []
        current = start_time
        while current < end_time:
            time_key = _format_time_label(current, time_range)
            if time_key in time_grouped_packet_loss:
                # 计算该时间段的平均丢包率
                avg_loss = sum(time_grouped_packet_loss[time_key]) / len(time_grouped_packet_loss[time_key])
                packet_loss_time_series.append(round(avg_loss, 2))
            else:
                packet_loss_time_series.append(0)
            current += interval
        
        # 计算整体平均丢包率
        avg_packet_loss = round(sum(packet_losses) / len(packet_losses), 2) if packet_losses else 0
        min_packet_loss = round(min(packet_losses), 2) if packet_losses else 0
        max_packet_loss = round(max(packet_losses), 2) if packet_losses else 0
        
        return {
            'response_time_stats': {
                'tcp_connect': calc_stats(tcp_times),
                'ping_latency': calc_stats(ping_times),
                'http_response': calc_stats(http_times),
                'api_response': calc_stats(api_times)
            },
            'packet_loss_stats': {
                'ping_packet_loss': [min_packet_loss, avg_packet_loss, max_packet_loss],
                'ping_packet_loss_data': packet_loss_time_series  # 新增时间序列数据
            }
        }
        
    except Exception as e:
        print(f'Error calculating performance stats: {e}')
        print(traceback.format_exc())
        # 返回默认值
        return {
            'response_time_stats': {
                'tcp_connect': [0, 0, 0],
                'ping_latency': [0, 0, 0],
                'http_response': [0, 0, 0],
                'api_response': [0, 0, 0]
            },
            'packet_loss_stats': {
                'ping_packet_loss': [0, 0, 0],
                'ping_packet_loss_data': []
            }
        }


@v2_bp.route('/reports/overview', methods=['GET'])
@token_required
@tenant_required
def get_report_overview_v2():
    """
    获取报表总览数据 - v2（强制租户隔离）
    返回结构与v1基本一致，并新增按任务类型的成功率趋势 type_success_trends 以便前端多折线展示
    """
    try:
        time_range = request.args.get('time_range', '1d', type=str)
        start_time, end_time, interval = _calc_time_range(time_range)

        tenant_id = TenantContext.get_current_tenant_id()
        if not tenant_id:
            return jsonify({'code': 403, 'message': '缺少租户上下文'}), 403

        # 仅查询当前租户的数据
        base_query = db.session.query(Result, Task).join(Task, Result.task_id == Task.id).filter(
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )

        # 总体统计
        total_results = base_query.count()
        success_results = base_query.filter(Result.status == 'success').count()
        overall_success_rate = (success_results / total_results * 100.0) if total_results > 0 else 0.0

        # 按任务类型统计成功率
        type_rows = db.session.query(
            Task.type.label('type'),
            func.count(Result.id).label('total')
        ).join(Result, Task.id == Result.task_id).filter(
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        ).group_by(Task.type).all()

        task_type_data = []
        for row in type_rows:
            total = row.total or 0
            # 单独查询成功数量
            success = db.session.query(func.count(Result.id)).join(Task, Result.task_id == Task.id).filter(
                Task.tenant_id == tenant_id,
                Task.type == row.type,
                Result.status == 'success',
                Result.created_at >= start_time,
                Result.created_at <= end_time
            ).scalar() or 0
            rate = (success / total * 100.0) if total > 0 else 0.0
            task_type_data.append({
                'type': row.type,
                'total': int(total),
                'success': int(success),
                'success_rate': round(rate, 2)
            })

        # 成功率趋势（总体）
        trend_data = []
        current = start_time
        while current < end_time:
            nxt = current + interval
            period_total = base_query.filter(Result.created_at >= current, Result.created_at < nxt).count()
            period_success = base_query.filter(
                Result.created_at >= current, Result.created_at < nxt, Result.status == 'success'
            ).count()
            rate = (period_success / period_total * 100.0) if period_total > 0 else 0.0
            trend_data.append({
                'time': _format_time_label(current, time_range),
                'success_rate': round(rate, 2),
                'total': int(period_total),
                'success': int(period_success)
            })
            current = nxt

        # 各任务类型的成功率趋势
        type_success_trends = {t: [] for t in ['tcp', 'ping', 'http', 'api']}
        current = start_time
        while current < end_time:
            nxt = current + interval
            label = _format_time_label(current, time_range)
            # 预填零值
            for t in type_success_trends.keys():
                type_success_trends[t].append({'time': label, 'success_rate': 0.0, 'total': 0, 'success': 0})

            # 统计当前时间段按类型的成功率
            rows = db.session.query(
                Task.type.label('type'),
                func.count(Result.id).label('total')
            ).join(Result, Task.id == Result.task_id).filter(
                Task.tenant_id == tenant_id,
                Result.created_at >= current,
                Result.created_at < nxt
            ).group_by(Task.type).all()

            row_map = {r.type: r for r in rows}
            for t in type_success_trends.keys():
                if t in row_map:
                    total = row_map[t].total or 0
                    # 单独查询成功数量
                    success = db.session.query(func.count(Result.id)).join(Task, Result.task_id == Task.id).filter(
                        Task.tenant_id == tenant_id,
                        Task.type == t,
                        Result.status == 'success',
                        Result.created_at >= current,
                        Result.created_at < nxt
                    ).scalar() or 0
                    rate = (success / total * 100.0) if total > 0 else 0.0
                    # 替换同时间位的条目
                    type_success_trends[t][-1] = {
                        'time': label,
                        'success_rate': round(rate, 2),
                        'total': int(total),
                        'success': int(success)
                    }
            current = nxt

        # TOP10任务（成功率高）
        top_rows = db.session.query(
            Task.id, Task.name, Task.type, Task.target,
            func.count(Result.id).label('total')
        ).join(Result, Task.id == Result.task_id).filter(
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        ).group_by(Task.id, Task.name, Task.type, Task.target).having(
            func.count(Result.id) >= 10
        ).all()

        top_tasks_data = []
        for r in top_rows:
            # 单独查询成功数量
            success = db.session.query(func.count(Result.id)).filter(
                Result.task_id == r.id,
                Result.status == 'success',
                Result.created_at >= start_time,
                Result.created_at <= end_time
            ).scalar() or 0
            rate = (success / r.total * 100.0) if r.total > 0 else 0.0
            top_tasks_data.append({
                'task_id': r.id,
                'task_name': r.name,
                'task_type': r.type,
                'target': r.target,
                'total': int(r.total),
                'success': int(success),
                'success_rate': round(rate, 2)
            })
        
        # 按成功率排序
        top_tasks_data.sort(key=lambda x: x['success_rate'], reverse=True)
        top_tasks_data = top_tasks_data[:10]

        # 失败率最高任务（TOP10）
        worst_rows = db.session.query(
            Task.id, Task.name, Task.type, Task.target,
            func.count(Result.id).label('total')
        ).join(Result, Task.id == Result.task_id).filter(
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        ).group_by(Task.id, Task.name, Task.type, Task.target).having(
            func.count(Result.id) >= 10
        ).all()

        worst_tasks_data = []
        for r in worst_rows:
            # 单独查询成功数量
            success = db.session.query(func.count(Result.id)).filter(
                Result.task_id == r.id,
                Result.status == 'success',
                Result.created_at >= start_time,
                Result.created_at <= end_time
            ).scalar() or 0
            rate = (success / r.total * 100.0) if r.total > 0 else 0.0
            worst_tasks_data.append({
                'task_id': r.id,
                'task_name': r.name,
                'task_type': r.type,
                'target': r.target,
                'total': int(r.total),
                'success': int(success),
                'success_rate': round(rate, 2),
                'failure_rate': round(100.0 - rate, 2)
            })
        
        # 按成功率升序排序（失败率最高）
        worst_tasks_data.sort(key=lambda x: x['success_rate'])
        worst_tasks_data = worst_tasks_data[:10]

        # 计算性能指标统计
        performance_stats = _calculate_performance_stats(tenant_id, start_time, end_time)
        
        return jsonify({
            'code': 0,
            'data': {
                'overview': {
                    'total_results': int(total_results),
                    'success_results': int(success_results),
                    'overall_success_rate': round(overall_success_rate, 2)
                },
                'task_type_stats': task_type_data,
                'success_rate_trend': trend_data,
                'type_success_trends': type_success_trends,
                'top_tasks': top_tasks_data,
                'worst_tasks': worst_tasks_data,
                'performance_stats': performance_stats
            },
            'message': 'success'
        })

    except Exception as e:
        print('Error in get_report_overview_v2:', e)
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取报表总览数据失败: {str(e)}'
        }), 500


def _generate_port_analysis(tenant_id: str, start_time: datetime, end_time: datetime, task_id: int = None, port: str = None):
    """生成端口分析数据"""
    try:
        # 构建基础查询
        base_query = db.session.query(Result, Task).join(Task, Result.task_id == Task.id).filter(
            Task.type == 'tcp',
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        if port:
            base_query = base_query.filter(Task.target.like(f'%:{port}'))
        
        # 调试信息
        all_results = base_query.all()
        print(f"Debug TCP Port Analysis: Found {len(all_results)} results for tenant {tenant_id}")
        for result, task in all_results[:5]:  # 只打印前5个结果
            print(f"Debug: Task {task.name} ({task.target}) - Status: {result.status}, Details: {result.details}")
        
        # 按端口统计数据
        port_stats = {}
        error_stats = {}
        
        for result, task in base_query.all():
            # 提取端口号
            target_port = 'unknown'
            if ':' in task.target:
                target_port = task.target.split(':')[-1]
            
            if target_port not in port_stats:
                port_stats[target_port] = {
                    'total': 0,
                    'success': 0,
                    'response_times': []
                }
            
            port_stats[target_port]['total'] += 1
            
            if result.status == 'success':
                port_stats[target_port]['success'] += 1
                
                # 提取响应时间
                if result.details:
                    try:
                        details = json.loads(result.details) if isinstance(result.details, str) else result.details
                        connect_time = details.get('connect_time') or details.get('response_time') or result.response_time
                        if connect_time and connect_time > 0:
                            port_stats[target_port]['response_times'].append(float(connect_time))
                    except (json.JSONDecodeError, TypeError):
                        pass
            else:
                # 统计错误类型
                error_message = 'Connection Failed'
                if result.details:
                    try:
                        details = json.loads(result.details) if isinstance(result.details, str) else result.details
                        error_message = details.get('message', 'Connection Failed')
                    except (json.JSONDecodeError, TypeError):
                        pass
                
                error_type = error_message
                if 'timeout' in error_type.lower():
                    error_type = 'Connection Timeout'
                elif 'refused' in error_type.lower():
                    error_type = 'Connection Refused'
                elif 'unreachable' in error_type.lower():
                    error_type = 'Host Unreachable'
                else:
                    error_type = 'Other Error'
                
                if error_type not in error_stats:
                    error_stats[error_type] = 0
                error_stats[error_type] += 1
        
        # 生成端口成功率数据
        port_success = []
        for port_num, stats in port_stats.items():
            success_rate = (stats['success'] / stats['total'] * 100) if stats['total'] > 0 else 0
            port_success.append({
                'port': port_num,
                'success_rate': round(success_rate, 2),
                'total': stats['total'],
                'success': stats['success']
            })
        
        # 生成端口响应时间数据
        port_response = []
        for port_num, stats in port_stats.items():
            if stats['response_times']:
                avg_time = sum(stats['response_times']) / len(stats['response_times'])
                max_time = max(stats['response_times'])
                port_response.append({
                    'port': port_num,
                    'avg_time': round(avg_time, 2),
                    'max_time': round(max_time, 2),
                    'min_time': round(min(stats['response_times']), 2)
                })
            else:
                port_response.append({
                    'port': port_num,
                    'avg_time': 0,
                    'max_time': 0,
                    'min_time': 0
                })
        
        # 生成错误分析数据
        error_analysis = []
        total_errors = sum(error_stats.values())
        for error_type, count in error_stats.items():
            percentage = (count / total_errors * 100) if total_errors > 0 else 0
            error_analysis.append({
                'name': error_type,
                'value': count,
                'percentage': round(percentage, 2)
            })
        
        return {
            'port_success': port_success,
            'port_response': port_response,
            'error_analysis': error_analysis
        }
        
    except Exception as e:
        print(f'Error generating port analysis: {str(e)}')
        return {
            'port_success': [],
            'port_response': [],
            'error_analysis': []
        }


def _generate_connection_time_distribution(connection_times: list):
    """生成连接时间分布数据"""
    try:
        if not connection_times:
            return []
        
        # 定义时间范围
        ranges = [
            {'range': '0-50ms', 'min': 0, 'max': 50},
            {'range': '50-100ms', 'min': 50, 'max': 100},
            {'range': '100-200ms', 'min': 100, 'max': 200},
            {'range': '200-500ms', 'min': 200, 'max': 500},
            {'range': '500ms+', 'min': 500, 'max': float('inf')}
        ]
        
        distribution = []
        for range_info in ranges:
            count = sum(1 for time in connection_times 
                       if range_info['min'] <= time < range_info['max'])
            distribution.append({
                'range': range_info['range'],
                'count': count,
                'percentage': round((count / len(connection_times) * 100), 2) if connection_times else 0
            })
        
        return distribution
        
    except Exception as e:
        print(f'Error generating connection time distribution: {str(e)}')
        return []


@v2_bp.route('/reports/<int:report_id>/export', methods=['POST'])
@token_required
@tenant_required
def export_report_v2(report_id):
    """导出报表 - v2（支持租户隔离）"""
    try:
        data = request.get_json()
        export_format = data.get('format', 'excel')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        filters = data.get('filters', {})
        
        tenant_id = TenantContext.get_current_tenant_id()
        if not tenant_id:
            return jsonify({'code': 403, 'message': '缺少租户上下文'}), 403
        
        # 根据report_id确定报表类型
        report_type_map = {
            1: 'overview',
            2: 'tcp', 
            3: 'ping',
            4: 'http',
            5: 'api'
        }
        report_type = report_type_map.get(report_id, 'overview')
        
        # 解析日期范围
        if start_date:
            start_time = datetime.strptime(start_date, '%Y-%m-%d')
        else:
            start_time = datetime.now() - timedelta(days=7)
            
        if end_date:
            end_time = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        else:
            end_time = datetime.now()
        
        # 生成Excel文件
        if export_format == 'excel':
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f'{report_type.upper()}报表'
            
            # 设置标题
            title = f'{report_type.upper()}专项报表'
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:F1')
            
            # 设置时间范围
            time_range = f'时间范围: {start_time.strftime("%Y-%m-%d")} 至 {end_time.strftime("%Y-%m-%d")}'
            ws['A2'] = time_range
            ws['A2'].font = Font(size=12)
            ws.merge_cells('A2:F2')
            
            # 设置表头
            headers = ['时间', '总数', '成功数', '失败数', '成功率(%)', '平均响应时间(ms)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # 查询实际数据（按小时统计）
            base_query = db.session.query(Result, Task).join(Task, Result.task_id == Task.id).filter(
                Task.tenant_id == tenant_id,
                Result.created_at >= start_time,
                Result.created_at < end_time
            )
            
            # 如果指定了任务类型，添加过滤条件
            if report_type != 'overview':
                base_query = base_query.filter(Task.type == report_type)
            
            # 按小时分组统计数据
            current_time = start_time
            row_idx = 5
            
            while current_time < end_time:
                next_time = current_time + timedelta(hours=1)
                
                # 查询该小时的数据
                hour_query = base_query.filter(
                    Result.created_at >= current_time,
                    Result.created_at < next_time
                )
                
                total_count = hour_query.count()
                success_count = hour_query.filter(Result.status == 'success').count()
                failure_count = total_count - success_count
                success_rate = (success_count / total_count * 100.0) if total_count > 0 else 0.0
                
                # 计算平均响应时间
                avg_response_time = 0.0
                if total_count > 0:
                    response_times = [r.Result.response_time for r in hour_query.all() if r.Result.response_time is not None]
                    if response_times:
                        avg_response_time = sum(response_times) / len(response_times)
                
                # 写入数据
                row_data = [
                    current_time.strftime('%Y-%m-%d %H:%M'),
                    total_count,
                    success_count,
                    failure_count,
                    round(success_rate, 2),
                    round(avg_response_time, 2)
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
                current_time = next_time
                row_idx += 1
            
            # 调整列宽
            column_letters = ['A', 'B', 'C', 'D', 'E', 'F']
            for i, letter in enumerate(column_letters):
                max_length = 0
                for row in ws.iter_rows(min_col=i+1, max_col=i+1):
                    for cell in row:
                        if hasattr(cell, 'value') and cell.value is not None:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                adjusted_width = min(max(max_length + 2, 12), 25)
                ws.column_dimensions[letter].width = adjusted_width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 生成文件名
            filename = f'{report_type}_report_{start_time.strftime("%Y%m%d")}_{end_time.strftime("%Y%m%d")}.xlsx'
            
            # 返回文件
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        else:
            # 暂不支持其他格式
            return jsonify({
                'code': 1,
                'data': None,
                'message': '暂不支持该导出格式'
            }), 400
        
    except Exception as e:
        print(f"Error in export_report_v2: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'导出报表失败: {str(e)}'
        }), 500


def _format_time_label(time_obj, time_range):
    """格式化时间标签"""
    if time_range == '1h':
        return time_obj.strftime('%H:%M')
    elif time_range == '1d':
        return time_obj.strftime('%H:%M')
    elif time_range == '7d':
        return time_obj.strftime('%m-%d')
    elif time_range == '30d':
        return time_obj.strftime('%m-%d')
    else:
        return time_obj.strftime('%Y-%m-%d %H:%M')


@v2_bp.route('/reports/api', methods=['GET'])
@token_required
@tenant_required
def get_api_report_v2():
    """获取API专项报表数据 - v2版本（强制租户隔离）"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        task_id = request.args.get('task_id', type=int)
        transaction_type = request.args.get('transaction_type', type=str)
        
        # 计算时间范围
        start_time, end_time, interval = _calc_time_range(time_range)
        
        # 获取当前租户ID
        tenant_id = TenantContext.get_current_tenant_id()
        if not tenant_id:
            return jsonify({'code': 403, 'message': '缺少租户上下文'}), 403
        
        # 构建基础查询（仅当前租户）
        base_query = db.session.query(Result, Task).join(Task, Result.task_id == Task.id).filter(
            Task.type == 'api',
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        # API关键指标统计
        total_transactions = base_query.count()
        successful_transactions = base_query.filter(Result.status == 'success').count()
        transaction_success_rate = (successful_transactions / total_transactions * 100) if total_transactions > 0 else 0
        
        # 计算断言通过率（从details中解析）
        assertion_passed = 0
        total_with_assertions = 0
        avg_transaction_time = 0
        
        for result, task in base_query.all():
            if result.details:
                try:
                    details = json.loads(result.details) if isinstance(result.details, str) else result.details
                    
                    # 检查断言结果
                    if 'assertions' in details or 'assertion_results' in details:
                        total_with_assertions += 1
                        assertions = details.get('assertions', details.get('assertion_results', []))
                        if isinstance(assertions, list) and all(a.get('passed', False) for a in assertions):
                            assertion_passed += 1
                        elif isinstance(assertions, dict) and assertions.get('all_passed', False):
                            assertion_passed += 1
                except (json.JSONDecodeError, TypeError):
                    pass
        
        assertion_pass_rate = (assertion_passed / total_with_assertions * 100) if total_with_assertions > 0 else 0
        
        # 计算平均事务时间
        response_times = [r.Result.response_time for r in base_query.all() if r.Result.response_time is not None]
        avg_transaction_time = sum(response_times) / len(response_times) if response_times else 0
        
        # 事务成功率趋势数据
        trend_data = []
        current = start_time
        while current < end_time:
            next_time = current + interval
            
            period_query = base_query.filter(
                Result.created_at >= current,
                Result.created_at < next_time
            )
            
            period_total = period_query.count()
            period_success = period_query.filter(Result.status == 'success').count()
            period_rate = (period_success / period_total * 100) if period_total > 0 else 0
            
            trend_data.append({
                'time': _format_time_label(current, time_range),
                'success_rate': round(period_rate, 2),
                'total': period_total,
                'success': period_success
            })
            
            current = next_time
        
        # 获取API任务列表
        api_tasks = db.session.query(Task).filter(
            Task.type == 'api',
            Task.tenant_id == tenant_id
        ).all()
        
        task_list = []
        for task in api_tasks:
            task_results = db.session.query(Result).filter(
                Result.task_id == task.id,
                Result.created_at >= start_time,
                Result.created_at <= end_time
            )
            
            task_total = task_results.count()
            task_success = task_results.filter(Result.status == 'success').count()
            task_success_rate = (task_success / task_total * 100) if task_total > 0 else 0
            
            # 计算任务的性能分析数据
            task_response_times = []
            for result in task_results.all():
                if result.response_time is not None:
                    task_response_times.append(result.response_time)
                elif result.details:
                    try:
                        details = json.loads(result.details) if isinstance(result.details, str) else result.details
                        if 'response_time' in details and details['response_time'] is not None:
                            task_response_times.append(details['response_time'])
                    except (json.JSONDecodeError, TypeError):
                        pass
            
            task_avg_time = sum(task_response_times) / len(task_response_times) if task_response_times else 0
            
            # 计算P95和P99响应时间
            task_response_times_sorted = sorted(task_response_times) if task_response_times else [0]
            p95_index = int(len(task_response_times_sorted) * 0.95) if task_response_times_sorted else 0
            p99_index = int(len(task_response_times_sorted) * 0.99) if task_response_times_sorted else 0
            p95_response_time = task_response_times_sorted[p95_index] if p95_index < len(task_response_times_sorted) else task_avg_time * 1.5
            p99_response_time = task_response_times_sorted[p99_index] if p99_index < len(task_response_times_sorted) else task_avg_time * 2
            
            # 计算任务级别的断言通过率
            task_assertion_passed = 0
            task_assertion_total = 0
            step_breakdown_data = []
            
            for result in task_results.all():
                if result.details:
                    try:
                        details = json.loads(result.details) if isinstance(result.details, str) else result.details
                        
                        # 统计断言通过
                        if 'assertions' in details or 'assertion_results' in details:
                            task_assertion_total += 1
                            assertions = details.get('assertions', details.get('assertion_results', []))
                            if isinstance(assertions, list) and all(a.get('passed', False) for a in assertions):
                                task_assertion_passed += 1
                            elif isinstance(assertions, dict) and assertions.get('all_passed', False):
                                task_assertion_passed += 1
                        
                        # 收集步骤分解数据
                        print(f"Debug: Processing result for task {task.name}, details structure: {type(details)}")
                        # 检查 details.details.steps 结构
                        if 'details' in details and isinstance(details['details'], dict):
                            inner_details = details['details']
                            print(f"Debug: Found inner_details, checking for steps: {'steps' in inner_details}")
                            if 'steps' in inner_details and isinstance(inner_details['steps'], list):
                                print(f"Debug: Found {len(inner_details['steps'])} steps in inner_details")
                                for step in inner_details['steps']:
                                    if isinstance(step, dict) and 'name' in step and 'response_time' in step:
                                        print(f"Debug: Adding step {step['name']} with time {step['response_time']}")
                                        step_breakdown_data.append({
                                            'name': step['name'],
                                            'time': step['response_time']
                                        })
                        # 也检查直接的 details.steps 结构（向后兼容）
                        elif 'steps' in details and isinstance(details['steps'], list):
                            print(f"Debug: Found {len(details['steps'])} steps in direct details")
                            for step in details['steps']:
                                if isinstance(step, dict) and 'name' in step and 'response_time' in step:
                                    print(f"Debug: Adding step {step['name']} with time {step['response_time']}")
                                    step_breakdown_data.append({
                                        'name': step['name'],
                                        'time': step['response_time']
                                    })
                    except (json.JSONDecodeError, TypeError):
                        pass
            
            task_assertion_pass_rate = (task_assertion_passed / task_assertion_total * 100) if task_assertion_total > 0 else 0
            
            # 处理步骤分解数据
            step_summary = []
            print(f"Debug: Task {task.name} collected step_breakdown_data: {step_breakdown_data}")
            if step_breakdown_data:
                step_dict = {}
                for step in step_breakdown_data:
                    if isinstance(step, dict) and 'name' in step and 'time' in step:
                        step_name = step['name']
                        step_time = step['time']
                        if step_name not in step_dict:
                            step_dict[step_name] = []
                        step_dict[step_name].append(step_time)
                
                for step_name, times in step_dict.items():
                    avg_time = sum(times) / len(times) if times else 0
                    step_summary.append({
                        'name': step_name,
                        'avg_time': round(avg_time, 2)
                    })
            print(f"Debug: Task {task.name} final step_summary: {step_summary}")
            
            # 获取最后执行时间
            last_result = task_results.order_by(Result.created_at.desc()).first()
            last_execution = last_result.created_at.strftime('%Y-%m-%d %H:%M:%S') if last_result else 'N/A'
            
            # 安全处理task.config，确保它是字典类型
            config = {}
            if task.config:
                if isinstance(task.config, dict):
                    config = task.config
                elif isinstance(task.config, str):
                    try:
                        config = json.loads(task.config)
                    except (json.JSONDecodeError, TypeError):
                        config = {}
            
            task_list.append({
                'task_id': task.id,
                'task_name': task.name,
                'transaction_type': config.get('transaction_type', 'API事务'),
                'step_count': len(config.get('steps', [])),
                'success_rate': round(task_success_rate, 2),
                'assertion_pass_rate': round(task_assertion_pass_rate, 2),
                'avg_response_time': round(task_avg_time, 2),
                'p95_response_time': round(p95_response_time, 2),
                'p99_response_time': round(p99_response_time, 2),
                'step_breakdown': step_summary,
                'total_executions': task_total,
                'last_execution_time': last_execution
            })
        
        # 失败原因分析
        failure_reasons = {
            '断言失败': 0,
            '接口超时': 0,
            '服务器错误': 0,
            '参数错误': 0,
            '网络异常': 0,
            '其他错误': 0
        }
        
        failed_results = base_query.filter(Result.status != 'success').all()
        for result, task in failed_results:
            if result.details:
                try:
                    details = json.loads(result.details) if isinstance(result.details, str) else result.details
                    if 'error_type' in details:
                        error_type = details['error_type']
                        if error_type in failure_reasons:
                            failure_reasons[error_type] += 1
                        else:
                            failure_reasons['其他错误'] += 1
                    else:
                        failure_reasons['其他错误'] += 1
                except:
                    failure_reasons['其他错误'] += 1
        
        failure_analysis = []
        total_failures = sum(failure_reasons.values())
        for reason, count in failure_reasons.items():
            percentage = (count / total_failures * 100) if total_failures > 0 else 0
            failure_analysis.append({
                'reason': reason,
                'count': count,
                'percentage': round(percentage, 2)
            })
        
        # 生成性能分析数据
        performance_analysis = _generate_performance_analysis_v2(task_list)
        
        return jsonify({
            'code': 0,
            'data': {
                'metrics': {
                    'transaction_success_rate': round(transaction_success_rate, 2),
                    'assertion_pass_rate': round(assertion_pass_rate, 2),
                    'avg_transaction_time': round(avg_transaction_time, 2),
                    'failure_rate': round(100 - transaction_success_rate, 2)
                },
                'success_rate_trend': trend_data,
                'task_list': task_list,
                'failure_reasons': failure_analysis,
                'transaction_success_trend': trend_data,
                'failure_analysis': failure_analysis,
                'task_details': task_list,
                'performance_analysis': performance_analysis,
                'total_tasks': len(task_list)
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f'Error in get_api_report_v2: {str(e)}')
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取API专项报表数据失败: {str(e)}'
        }), 500


def _generate_performance_analysis_v2(task_list):
    """生成API性能分析数据 - v2版本"""
    if not task_list:
        return {
            'step_breakdown': [],
            'transaction_performance': [],
            'performance_summary': {
                'avg_response_time': 0,
                'p95_response_time': 0,
                'p99_response_time': 0,
                'total_transactions': 0
            }
        }
    
    # 步骤分解数据
    step_breakdown_data = []
    transaction_performance_data = []
    
    total_avg_time = 0
    total_p95_time = 0
    total_p99_time = 0
    total_transactions = 0
    
    for task in task_list:
        # 步骤分解数据
        print(f"Debug: Task {task['task_name']} step_breakdown: {task.get('step_breakdown')}")
        if task.get('step_breakdown'):
            step_breakdown_data.append({
                'task_name': task['task_name'],
                'task_id': task['task_id'],
                'steps': task['step_breakdown']
            })
        
        # 事务性能对比数据
        transaction_performance_data.append({
            'task_name': task['task_name'],
            'task_id': task['task_id'],
            'avg_response_time': task.get('avg_response_time', 0),
            'p95_response_time': task.get('p95_response_time', 0),
            'p99_response_time': task.get('p99_response_time', 0),
            'success_rate': task.get('success_rate', 0),
            'assertion_pass_rate': task.get('assertion_pass_rate', 0),
            'total_requests': task.get('total_executions', 0)
        })
        
        # 累计统计
        total_executions = task.get('total_executions', 0)
        total_avg_time += task.get('avg_response_time', 0) * total_executions
        total_p95_time += task.get('p95_response_time', 0) * total_executions
        total_p99_time += task.get('p99_response_time', 0) * total_executions
        total_transactions += total_executions
    
    # 计算总体性能指标
    performance_summary = {
        'avg_response_time': round(total_avg_time / total_transactions, 2) if total_transactions > 0 else 0,
        'p95_response_time': round(total_p95_time / total_transactions, 2) if total_transactions > 0 else 0,
        'p99_response_time': round(total_p99_time / total_transactions, 2) if total_transactions > 0 else 0,
        'total_transactions': total_transactions
    }
    
    return {
        'step_breakdown': step_breakdown_data,
        'transaction_performance': transaction_performance_data,
        'performance_summary': performance_summary
    }


@v2_bp.route('/reports/http', methods=['GET'])
@token_required
@tenant_required
def get_http_report_v2():
    """获取HTTP专项报表数据 - v2版本（强制租户隔离）"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        task_id = request.args.get('task_id', type=int)
        status_code = request.args.get('status_code', type=str)
        
        # 计算时间范围
        start_time, end_time, interval = _calc_time_range(time_range)
        
        # 获取当前租户ID
        tenant_id = TenantContext.get_current_tenant_id()
        if not tenant_id:
            return jsonify({'code': 403, 'message': '缺少租户上下文'}), 403
        
        # 构建基础查询（仅当前租户）
        base_query = db.session.query(Result, Task).join(Task, Result.task_id == Task.id).filter(
            Task.type == 'http',
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        # HTTP关键指标统计
        total_requests = base_query.count()
        successful_requests = base_query.filter(Result.status == 'success').count()
        success_rate = (successful_requests / total_requests * 100) if total_requests > 0 else 0
        
        # 计算平均响应时间
        response_times = [r.Result.response_time for r in base_query.all() if r.Result.response_time is not None]
        avg_response_time = sum(response_times) / len(response_times) if response_times else 0
        
        # 响应时间趋势数据
        trend_data = []
        current = start_time
        while current < end_time:
            next_time = current + interval
            
            period_query = base_query.filter(
                Result.created_at >= current,
                Result.created_at < next_time
            )
            
            period_total = period_query.count()
            period_success = period_query.filter(Result.status == 'success').count()
            period_rate = (period_success / period_total * 100) if period_total > 0 else 0
            
            # 计算该时段的平均响应时间
            period_times = [r.Result.response_time for r in period_query.all() if r.Result.response_time is not None]
            period_avg_time = sum(period_times) / len(period_times) if period_times else 0
            
            trend_data.append({
                'time': _format_time_label(current, time_range),
                'success_rate': round(period_rate, 2),
                'avg_response_time': round(period_avg_time, 2),
                'total': period_total,
                'success': period_success
            })
            
            current = next_time
        
        # 获取HTTP任务列表
        http_tasks = db.session.query(Task).filter(
            Task.type == 'http',
            Task.tenant_id == tenant_id
        ).all()
        
        task_list = []
        for task in http_tasks:
            task_results = db.session.query(Result).filter(
                Result.task_id == task.id,
                Result.created_at >= start_time,
                Result.created_at <= end_time
            )
            
            task_total = task_results.count()
            task_success = task_results.filter(Result.status == 'success').count()
            task_success_rate = (task_success / task_total * 100) if task_total > 0 else 0
            
            # 计算任务的平均响应时间
            task_response_times = [r.response_time for r in task_results.all() if r.response_time is not None]
            task_avg_time = sum(task_response_times) / len(task_response_times) if task_response_times else 0
            
            # 获取最后执行时间
            last_result = task_results.order_by(Result.created_at.desc()).first()
            last_execution = last_result.created_at.strftime('%Y-%m-%d %H:%M:%S') if last_result else 'N/A'
            
            task_list.append({
                'task_id': task.id,
                'task_name': task.name,
                'target_url': task.target,
                'success_rate': round(task_success_rate, 2),
                'avg_response_time': round(task_avg_time, 2),
                'total_requests': task_total,
                'last_execution_time': last_execution
            })
        
        return jsonify({
            'code': 0,
            'data': {
                'metrics': {
                    'success_rate': round(success_rate, 2),
                    'avg_response_time': round(avg_response_time, 2),
                    'total_requests': total_requests,
                    'failure_rate': round(100 - success_rate, 2)
                },
                'response_time_trend': trend_data,
                'task_list': task_list,
                'total_tasks': len(task_list)
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f'Error in get_http_report_v2: {str(e)}')
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取HTTP专项报表数据失败: {str(e)}'
        }), 500


@v2_bp.route('/reports/ping', methods=['GET'])
@token_required
@tenant_required
def get_ping_report_v2():
    """获取Ping专项报表数据 - v2版本（强制租户隔离）"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        task_id = request.args.get('task_id', type=int)
        location = request.args.get('location', type=str)
        
        # 计算时间范围
        start_time, end_time, interval = _calc_time_range(time_range)
        
        # 获取当前租户ID
        tenant_id = TenantContext.get_current_tenant_id()
        if not tenant_id:
            return jsonify({'code': 403, 'message': '缺少租户上下文'}), 403
        
        # 构建基础查询（仅当前租户）
        base_query = db.session.query(Result, Task).join(Task, Result.task_id == Task.id).filter(
            Task.type == 'ping',
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        if location:
            base_query = base_query.filter(Result.agent_area == location)
        
        # Ping关键指标统计
        total_pings = base_query.count()
        successful_pings = base_query.filter(Result.status == 'success').count()
        success_rate = (successful_pings / total_pings * 100) if total_pings > 0 else 0
        
        # 计算平均延迟和丢包率
        avg_latency = 0
        packet_loss_rate = 0
        
        latencies = []
        packet_losses = []
        
        for result, task in base_query.all():
            if result.details:
                try:
                    details = json.loads(result.details) if isinstance(result.details, str) else result.details
                    
                    # 提取延迟数据
                    rtt_avg = details.get('rtt_avg') or details.get('execution_time') or result.response_time
                    if rtt_avg and rtt_avg > 0:
                        latencies.append(float(rtt_avg))
                    
                    # 提取丢包率数据
                    packet_loss = details.get('packet_loss')
                    if packet_loss is not None:
                        packet_losses.append(float(packet_loss))
                        
                except (json.JSONDecodeError, TypeError):
                    pass
        
        avg_latency = sum(latencies) / len(latencies) if latencies else 0
        packet_loss_rate = sum(packet_losses) / len(packet_losses) if packet_losses else 0
        
        # 延迟趋势数据
        trend_data = []
        current = start_time
        while current < end_time:
            next_time = current + interval
            
            period_query = base_query.filter(
                Result.created_at >= current,
                Result.created_at < next_time
            )
            
            period_total = period_query.count()
            period_success = period_query.filter(Result.status == 'success').count()
            period_rate = (period_success / period_total * 100) if period_total > 0 else 0
            
            # 计算该时段的延迟统计
            period_latencies = []
            for result, task in period_query.all():
                if result.details:
                    try:
                        details = json.loads(result.details) if isinstance(result.details, str) else result.details
                        rtt_avg = details.get('rtt_avg') or details.get('execution_time') or result.response_time
                        if rtt_avg and rtt_avg > 0:
                            period_latencies.append(float(rtt_avg))
                    except (json.JSONDecodeError, TypeError):
                        pass
            
            period_avg_latency = sum(period_latencies) / len(period_latencies) if period_latencies else 0
            period_min_latency = min(period_latencies) if period_latencies else 0
            period_max_latency = max(period_latencies) if period_latencies else 0
            
            trend_data.append({
                'time': _format_time_label(current, time_range),
                'success_rate': round(period_rate, 2),
                'avg_latency': round(period_avg_latency, 2),
                'min_latency': round(period_min_latency, 2),
                'max_latency': round(period_max_latency, 2),
                'total': period_total,
                'success': period_success
            })
            
            current = next_time
        
        # 获取Ping任务列表
        ping_tasks = db.session.query(Task).filter(
            Task.type == 'ping',
            Task.tenant_id == tenant_id
        ).all()
        
        task_list = []
        for task in ping_tasks:
            task_results = db.session.query(Result).filter(
                Result.task_id == task.id,
                Result.created_at >= start_time,
                Result.created_at <= end_time
            )
            
            task_total = task_results.count()
            task_success = task_results.filter(Result.status == 'success').count()
            task_success_rate = (task_success / task_total * 100) if task_total > 0 else 0
            
            # 计算任务的平均延迟
            task_latencies = []
            for result in task_results.all():
                if result.details:
                    try:
                        details = json.loads(result.details) if isinstance(result.details, str) else result.details
                        rtt_avg = details.get('rtt_avg') or details.get('execution_time') or result.response_time
                        if rtt_avg and rtt_avg > 0:
                            task_latencies.append(float(rtt_avg))
                    except (json.JSONDecodeError, TypeError):
                        pass
            
            task_avg_latency = sum(task_latencies) / len(task_latencies) if task_latencies else 0
            
            # 获取最后执行时间
            last_result = task_results.order_by(Result.created_at.desc()).first()
            last_execution = last_result.created_at.strftime('%Y-%m-%d %H:%M:%S') if last_result else 'N/A'
            
            task_list.append({
                'task_id': task.id,
                'task_name': task.name,
                'target_host': task.target,
                'success_rate': round(task_success_rate, 2),
                'avg_latency': round(task_avg_latency, 2),
                'total_pings': task_total,
                'last_execution_time': last_execution
            })
        
        # 计算延迟分布统计
        latency_distribution = [
            {'range': '0-20ms', 'value': 0},
            {'range': '20-50ms', 'value': 0},
            {'range': '50-100ms', 'value': 0},
            {'range': '100-200ms', 'value': 0},
            {'range': '>200ms', 'value': 0}
        ]
        
        for latency in latencies:
            if latency <= 20:
                latency_distribution[0]['value'] += 1
            elif latency <= 50:
                latency_distribution[1]['value'] += 1
            elif latency <= 100:
                latency_distribution[2]['value'] += 1
            elif latency <= 200:
                latency_distribution[3]['value'] += 1
            else:
                latency_distribution[4]['value'] += 1
        
        # 计算抖动数据（基于延迟变化）
        jitter_data = []
        if len(latencies) > 1:
            for i in range(1, len(latencies)):
                jitter = abs(latencies[i] - latencies[i-1])
                jitter_data.append(jitter)
        
        avg_jitter = sum(jitter_data) / len(jitter_data) if jitter_data else 0
        
        # 地理位置分析数据（基于agent_area）
        geographic_data = {}
        for result, task in base_query.all():
            area = result.agent_area or '未知'
            if area not in geographic_data:
                geographic_data[area] = {'latencies': [], 'packet_losses': []}
            
            if result.details:
                try:
                    details = json.loads(result.details) if isinstance(result.details, str) else result.details
                    rtt_avg = details.get('rtt_avg') or details.get('execution_time') or result.response_time
                    if rtt_avg and rtt_avg > 0:
                        geographic_data[area]['latencies'].append(float(rtt_avg))
                    
                    packet_loss = details.get('packet_loss')
                    if packet_loss is not None:
                        geographic_data[area]['packet_losses'].append(float(packet_loss))
                except (json.JSONDecodeError, TypeError):
                    pass
        
        geographic_analysis = []
        for area, data in geographic_data.items():
            avg_area_latency = sum(data['latencies']) / len(data['latencies']) if data['latencies'] else 0
            avg_area_packet_loss = sum(data['packet_losses']) / len(data['packet_losses']) if data['packet_losses'] else 0
            geographic_analysis.append({
                'region': area,
                'avg_latency': round(avg_area_latency, 2),
                'packet_loss': round(avg_area_packet_loss, 2)
            })
        
        return jsonify({
            'code': 0,
            'data': {
                'metrics': {
                    'success_rate': round(success_rate, 2),
                    'avg_latency': round(avg_latency, 2),
                    'packet_loss_rate': round(packet_loss_rate, 2),
                    'total_pings': total_pings,
                    'jitter': round(avg_jitter, 2)
                },
                'latency_trend': trend_data,
                'latency_distribution': latency_distribution,
                'performance_metrics': {
                    'latency_distribution': latency_distribution,
                    'jitter_analysis': jitter_data[:50] if jitter_data else [],  # 限制数据量
                    'geographic_analysis': geographic_analysis
                },
                'task_list': task_list,
                'total_tasks': len(task_list)
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f'Error in get_ping_report_v2: {str(e)}')
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取Ping专项报表数据失败: {str(e)}'
        }), 500


@v2_bp.route('/reports/tcp', methods=['GET'])
@token_required
@tenant_required
def get_tcp_report_v2():
    """获取TCP专项报表数据 - v2版本（强制租户隔离）"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        task_id = request.args.get('task_id', type=int)
        port = request.args.get('port', type=str)
        
        # 计算时间范围
        start_time, end_time, interval = _calc_time_range(time_range)
        
        # 获取当前租户ID
        tenant_id = TenantContext.get_current_tenant_id()
        if not tenant_id:
            return jsonify({'code': 403, 'message': '缺少租户上下文'}), 403
        
        # 构建基础查询（仅当前租户）
        base_query = db.session.query(Result, Task).join(Task, Result.task_id == Task.id).filter(
            Task.type == 'tcp',
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        if port:
            base_query = base_query.filter(Task.target.like(f'%:{port}'))
        
        # TCP关键指标统计
        total_connections = base_query.count()
        successful_connections = base_query.filter(Result.status == 'success').count()
        success_rate = (successful_connections / total_connections * 100) if total_connections > 0 else 0
        
        # 计算平均连接时间
        connection_times = []
        for result, task in base_query.all():
            if result.details:
                try:
                    details = json.loads(result.details) if isinstance(result.details, str) else result.details
                    connect_time = details.get('connect_time') or details.get('response_time') or result.response_time
                    if connect_time and connect_time > 0:
                        connection_times.append(float(connect_time))
                except (json.JSONDecodeError, TypeError):
                    pass
        
        avg_connect_time = sum(connection_times) / len(connection_times) if connection_times else 0
        
        # 连接时间趋势数据
        trend_data = []
        current = start_time
        while current < end_time:
            next_time = current + interval
            
            period_query = base_query.filter(
                Result.created_at >= current,
                Result.created_at < next_time
            )
            
            period_total = period_query.count()
            period_success = period_query.filter(Result.status == 'success').count()
            period_rate = (period_success / period_total * 100) if period_total > 0 else 0
            
            # 计算该时段的平均连接时间
            period_times = []
            for result, task in period_query.all():
                if result.details:
                    try:
                        details = json.loads(result.details) if isinstance(result.details, str) else result.details
                        connect_time = details.get('connect_time') or details.get('response_time') or result.response_time
                        if connect_time and connect_time > 0:
                            period_times.append(float(connect_time))
                    except (json.JSONDecodeError, TypeError):
                        pass
            
            period_avg_time = sum(period_times) / len(period_times) if period_times else 0
            
            trend_data.append({
                'time': _format_time_label(current, time_range),
                'success_rate': round(period_rate, 2),
                'avg_connect_time': round(period_avg_time, 2),
                'total': period_total,
                'success': period_success
            })
            
            current = next_time
        
        # 获取TCP任务列表
        tcp_tasks = db.session.query(Task).filter(
            Task.type == 'tcp',
            Task.tenant_id == tenant_id
        ).all()
        
        task_list = []
        for task in tcp_tasks:
            task_results = db.session.query(Result).filter(
                Result.task_id == task.id,
                Result.created_at >= start_time,
                Result.created_at <= end_time
            )
            
            task_total = task_results.count()
            task_success = task_results.filter(Result.status == 'success').count()
            task_success_rate = (task_success / task_total * 100) if task_total > 0 else 0
            
            # 计算任务的平均连接时间
            task_times = []
            for result in task_results.all():
                if result.details:
                    try:
                        details = json.loads(result.details) if isinstance(result.details, str) else result.details
                        connect_time = details.get('connect_time') or details.get('response_time') or result.response_time
                        if connect_time and connect_time > 0:
                            task_times.append(float(connect_time))
                    except (json.JSONDecodeError, TypeError):
                        pass
            
            task_avg_time = sum(task_times) / len(task_times) if task_times else 0
            
            # 获取最后执行时间
            last_result = task_results.order_by(Result.created_at.desc()).first()
            last_execution = last_result.created_at.strftime('%Y-%m-%d %H:%M:%S') if last_result else 'N/A'
            
            task_list.append({
                'task_id': task.id,
                'task_name': task.name,
                'target_address': task.target,
                'success_rate': round(task_success_rate, 2),
                'avg_connect_time': round(task_avg_time, 2),
                'total_connections': task_total,
                'last_execution_time': last_execution
            })
        
        # 生成端口分析数据
        port_analysis = _generate_port_analysis(tenant_id, start_time, end_time, task_id, port)
        
        # 生成连接时间分布数据
        connection_time_distribution = _generate_connection_time_distribution(connection_times)
        
        return jsonify({
            'code': 0,
            'data': {
                'metrics': {
                    'success_rate': round(success_rate, 2),
                    'avg_connect_time': round(avg_connect_time, 2),
                    'total_connections': total_connections,
                    'failure_rate': round(100 - success_rate, 2)
                },
                'connect_time_trend': trend_data,
                'connection_time_distribution': connection_time_distribution,
                'port_analysis': port_analysis,
                'task_list': task_list,
                'total_tasks': len(task_list)
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f'Error in get_tcp_report_v2: {str(e)}')
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取TCP专项报表数据失败: {str(e)}'
        }), 500