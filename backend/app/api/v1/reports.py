from flask import request, jsonify, send_file, make_response
from . import bp
from app import db
from app.models.result import Result
from app.models.task import Task
from app.models.node import Node
from app.utils.auth_decorators import token_required
from app.utils.tenant_context import get_current_tenant_id, filter_by_tenant
from sqlalchemy import func, and_, or_
from datetime import datetime, timedelta
import json
import traceback
from dateutil import parser
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


@bp.route('/reports/overview', methods=['GET'])
@token_required
def get_report_overview():
    """获取报表总览数据"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        
        # 计算时间范围
        end_time = datetime.now()
        if time_range == '1h':
            start_time = end_time - timedelta(hours=1)
        elif time_range == '1d':
            start_time = end_time - timedelta(days=1)
        elif time_range == '7d':
            start_time = end_time - timedelta(days=7)
        elif time_range == '30d':
            start_time = end_time - timedelta(days=30)
        else:
            start_time = end_time - timedelta(days=1)
        
        # 获取总体统计数据（仅当前租户）
        tenant_id = get_current_tenant_id()
        total_results = filter_by_tenant(Result.query).join(Task).filter(
            Result.created_at >= start_time,
            Result.created_at <= end_time,
            Task.tenant_id == tenant_id
        ).count()
        
        success_results = filter_by_tenant(Result.query).join(Task).filter(
            Result.created_at >= start_time,
            Result.created_at <= end_time,
            Result.status == 'success',
            Task.tenant_id == tenant_id
        ).count()
        
        overall_success_rate = (success_results / total_results * 100) if total_results > 0 else 0
        
        # 按任务类型统计成功率
        task_type_stats = db.session.query(
            Task.type,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success')
        ).join(Result, Task.id == Result.task_id).filter(
            Result.created_at >= start_time,
            Result.created_at <= end_time
        ).group_by(Task.type).all()
        
        task_type_data = []
        for stat in task_type_stats:
            success_rate = (stat.success / stat.total * 100) if stat.total > 0 else 0
            task_type_data.append({
                'type': stat.type,
                'total': stat.total,
                'success': stat.success,
                'success_rate': round(success_rate, 2)
            })
        
        # 获取成功率趋势数据（按小时统计）
        trend_data = []
        current_time = start_time
        interval = timedelta(hours=1) if time_range in ['1h', '1d'] else timedelta(days=1)
        
        while current_time < end_time:
            next_time = current_time + interval
            
            period_total = Result.query.filter(
                Result.created_at >= current_time,
                Result.created_at < next_time
            ).count()
            
            period_success = Result.query.filter(
                Result.created_at >= current_time,
                Result.created_at < next_time,
                Result.status == 'success'
            ).count()
            
            period_success_rate = (period_success / period_total * 100) if period_total > 0 else 0
            
            trend_data.append({
                'time': current_time.strftime('%H:%M' if time_range in ['1h', '1d'] else '%m-%d'),
                'success_rate': round(period_success_rate, 2),
                'total': period_total,
                'success': period_success
            })
            
            current_time = next_time
        
        # 获取TOP10任务（按成功率排序）
        top_tasks = db.session.query(
            Task.id,
            Task.name,
            Task.type,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success')
        ).join(Result, Task.id == Result.task_id).filter(
            Result.created_at >= start_time,
            Result.created_at <= end_time
        ).group_by(Task.id, Task.name, Task.type).having(
            func.count(Result.id) >= 10  # 至少有10次执行记录
        ).order_by(
            (func.sum(func.case([(Result.status == 'success', 1)], else_=0)) * 100.0 / func.count(Result.id)).desc()
        ).limit(10).all()
        
        top_tasks_data = []
        for task in top_tasks:
            success_rate = (task.success / task.total * 100) if task.total > 0 else 0
            top_tasks_data.append({
                'task_id': task.id,
                'task_name': task.name,
                'task_type': task.type,
                'total': task.total,
                'success': task.success,
                'success_rate': round(success_rate, 2)
            })
        
        # 获取失败率最高的任务
        worst_tasks = db.session.query(
            Task.id,
            Task.name,
            Task.type,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success')
        ).join(Result, Task.id == Result.task_id).filter(
            Result.created_at >= start_time,
            Result.created_at <= end_time
        ).group_by(Task.id, Task.name, Task.type).having(
            func.count(Result.id) >= 10  # 至少有10次执行记录
        ).order_by(
            (func.sum(func.case([(Result.status == 'success', 1)], else_=0)) * 100.0 / func.count(Result.id)).asc()
        ).limit(10).all()
        
        worst_tasks_data = []
        for task in worst_tasks:
            success_rate = (task.success / task.total * 100) if task.total > 0 else 0
            worst_tasks_data.append({
                'task_id': task.id,
                'task_name': task.name,
                'task_type': task.type,
                'total': task.total,
                'success': task.success,
                'success_rate': round(success_rate, 2),
                'failure_rate': round(100 - success_rate, 2)
            })
        
        return jsonify({
            'code': 0,
            'data': {
                'overview': {
                    'total_results': total_results,
                    'success_results': success_results,
                    'overall_success_rate': round(overall_success_rate, 2)
                },
                'task_type_stats': task_type_data,
                'success_rate_trend': trend_data,
                'top_tasks': top_tasks_data,
                'worst_tasks': worst_tasks_data
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f"Error in get_report_overview: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取报表总览数据失败: {str(e)}'
        }), 500


@bp.route('/reports/tcp', methods=['GET'])
@token_required
def get_tcp_report():
    """获取TCP专项报表数据"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        task_id = request.args.get('task_id', type=int)
        port = request.args.get('port', type=str)
        
        # 计算时间范围
        end_time = datetime.now()
        if time_range == '1h':
            start_time = end_time - timedelta(hours=1)
        elif time_range == '1d':
            start_time = end_time - timedelta(days=1)
        elif time_range == '7d':
            start_time = end_time - timedelta(days=7)
        elif time_range == '30d':
            start_time = end_time - timedelta(days=30)
        else:
            start_time = end_time - timedelta(days=1)
        
        # 构建基础查询（仅当前租户）
        tenant_id = get_current_tenant_id()
        base_query = filter_by_tenant(Result.query).join(Task, Result.task_id == Task.id).filter(
            Task.type == 'tcp',
            Task.tenant_id == tenant_id,
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        if port:
            base_query = base_query.filter(Task.target.like(f'%:{port}'))
        
        # TCP关键指标统计
        total_connections = base_query.count()
        successful_connections = base_query.filter(Result.status == 'success').count()
        success_rate = (successful_connections / total_connections * 100) if total_connections > 0 else 0
        
        # 计算平均连接时间
        avg_connection_time = 0
        if total_connections > 0:
            connection_times = []
            results = base_query.all()
            for result in results:
                if result.result_data:
                    try:
                        data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                        if 'connection_time' in data:
                            connection_times.append(data['connection_time'])
                    except:
                        pass
            
            if connection_times:
                avg_connection_time = sum(connection_times) / len(connection_times)
        
        error_rate = 100 - success_rate
        
        # TCP连接成功率趋势
        trend_data = []
        current_time = start_time
        interval = timedelta(hours=1) if time_range in ['1h', '1d'] else timedelta(days=1)
        
        while current_time < end_time:
            next_time = current_time + interval
            
            period_query = base_query.filter(
                Result.created_at >= current_time,
                Result.created_at < next_time
            )
            
            period_total = period_query.count()
            period_success = period_query.filter(Result.status == 'success').count()
            period_success_rate = (period_success / period_total * 100) if period_total > 0 else 0
            
            trend_data.append({
                'time': current_time.strftime('%H:%M' if time_range in ['1h', '1d'] else '%m-%d'),
                'success_rate': round(period_success_rate, 2),
                'total': period_total,
                'success': period_success
            })
            
            current_time = next_time
        
        # 端口连通性分析
        port_stats = db.session.query(
            Task.target,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success')
        ).join(Result, Task.id == Result.task_id).filter(
            Task.type == 'tcp',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            port_stats = port_stats.filter(Task.id == task_id)
        
        port_stats = port_stats.group_by(Task.target).all()
        
        port_analysis = []
        for stat in port_stats:
            success_rate = (stat.success / stat.total * 100) if stat.total > 0 else 0
            # 提取端口号
            port_num = stat.target.split(':')[-1] if ':' in stat.target else 'unknown'
            port_analysis.append({
                'target': stat.target,
                'port': port_num,
                'total': stat.total,
                'success': stat.success,
                'success_rate': round(success_rate, 2)
            })
        
        # TCP任务详细统计
        task_stats = db.session.query(
            Task.id,
            Task.name,
            Task.target,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success'),
            func.max(Result.created_at).label('last_execution')
        ).join(Result, Task.id == Result.task_id).filter(
            Task.type == 'tcp',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            task_stats = task_stats.filter(Task.id == task_id)
        
        task_stats = task_stats.group_by(Task.id, Task.name, Task.target).all()
        
        task_details = []
        for stat in task_stats:
            success_rate = (stat.success / stat.total * 100) if stat.total > 0 else 0
            task_details.append({
                'task_id': stat.id,
                'task_name': stat.name,
                'target': stat.target,
                'total': stat.total,
                'success': stat.success,
                'success_rate': round(success_rate, 2),
                'last_execution': stat.last_execution.strftime('%Y-%m-%d %H:%M:%S') if stat.last_execution else None
            })
        
        return jsonify({
            'code': 0,
            'data': {
                'metrics': {
                    'success_rate': round(success_rate, 2),
                    'avg_connection_time': round(avg_connection_time, 2),
                    'total_connections': total_connections,
                    'error_rate': round(error_rate, 2)
                },
                'success_rate_trend': trend_data,
                'port_analysis': port_analysis,
                'task_details': task_details
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f"Error in get_tcp_report: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取TCP报表数据失败: {str(e)}'
        }), 500


@bp.route('/reports/ping', methods=['GET'])
@token_required
def get_ping_report():
    """获取Ping专项报表数据"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        task_id = request.args.get('task_id', type=int)
        location = request.args.get('location', type=str)
        
        # 计算时间范围
        end_time = datetime.now()
        if time_range == '1h':
            start_time = end_time - timedelta(hours=1)
        elif time_range == '1d':
            start_time = end_time - timedelta(days=1)
        elif time_range == '7d':
            start_time = end_time - timedelta(days=7)
        elif time_range == '30d':
            start_time = end_time - timedelta(days=30)
        else:
            start_time = end_time - timedelta(days=1)
        
        # 构建基础查询
        base_query = Result.query.join(Task, Result.task_id == Task.id).filter(
            Task.type == 'ping',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        if location:
            base_query = base_query.filter(Result.agent_area == location)
        
        # Ping关键指标统计
        total_pings = base_query.count()
        successful_pings = base_query.filter(Result.status == 'success').count()
        availability = (successful_pings / total_pings * 100) if total_pings > 0 else 0
        
        # 计算平均延迟、丢包率、抖动
        avg_latency = 0
        packet_loss_rate = 0
        avg_jitter = 0
        
        if total_pings > 0:
            latencies = []
            packet_losses = []
            jitters = []
            
            results = base_query.all()
            for result in results:
                if result.result_data:
                    try:
                        data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                        if 'avg_latency' in data:
                            latencies.append(data['avg_latency'])
                        if 'packet_loss' in data:
                            packet_losses.append(data['packet_loss'])
                        if 'jitter' in data:
                            jitters.append(data['jitter'])
                    except:
                        pass
            
            if latencies:
                avg_latency = sum(latencies) / len(latencies)
            if packet_losses:
                packet_loss_rate = sum(packet_losses) / len(packet_losses)
            if jitters:
                avg_jitter = sum(jitters) / len(jitters)
        
        # Ping延迟趋势
        trend_data = []
        current_time = start_time
        interval = timedelta(hours=1) if time_range in ['1h', '1d'] else timedelta(days=1)
        
        while current_time < end_time:
            next_time = current_time + interval
            
            period_results = base_query.filter(
                Result.created_at >= current_time,
                Result.created_at < next_time
            ).all()
            
            period_latencies = []
            period_packet_losses = []
            
            for result in period_results:
                if result.result_data:
                    try:
                        data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                        if 'avg_latency' in data:
                            period_latencies.append(data['avg_latency'])
                        if 'packet_loss' in data:
                            period_packet_losses.append(data['packet_loss'])
                    except:
                        pass
            
            period_avg_latency = sum(period_latencies) / len(period_latencies) if period_latencies else 0
            period_avg_packet_loss = sum(period_packet_losses) / len(period_packet_losses) if period_packet_losses else 0
            
            trend_data.append({
                'time': current_time.strftime('%H:%M' if time_range in ['1h', '1d'] else '%m-%d'),
                'avg_latency': round(period_avg_latency, 2),
                'packet_loss': round(period_avg_packet_loss, 2),
                'total': len(period_results)
            })
            
            current_time = next_time
        
        # 地理位置分析
        location_stats = db.session.query(
            Result.agent_area,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success')
        ).join(Task, Result.task_id == Task.id).filter(
            Task.type == 'ping',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            location_stats = location_stats.filter(Task.id == task_id)
        
        location_stats = location_stats.group_by(Result.agent_area).all()
        
        location_analysis = []
        for stat in location_stats:
            success_rate = (stat.success / stat.total * 100) if stat.total > 0 else 0
            location_analysis.append({
                'location': stat.agent_area or 'unknown',
                'total': stat.total,
                'success': stat.success,
                'success_rate': round(success_rate, 2)
            })
        
        # Ping任务详细统计
        task_stats = db.session.query(
            Task.id,
            Task.name,
            Task.target,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success'),
            func.max(Result.created_at).label('last_execution')
        ).join(Result, Task.id == Result.task_id).filter(
            Task.type == 'ping',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            task_stats = task_stats.filter(Task.id == task_id)
        
        task_stats = task_stats.group_by(Task.id, Task.name, Task.target).all()
        
        task_details = []
        for stat in task_stats:
            success_rate = (stat.success / stat.total * 100) if stat.total > 0 else 0
            task_details.append({
                'task_id': stat.id,
                'task_name': stat.name,
                'target': stat.target,
                'total': stat.total,
                'success': stat.success,
                'success_rate': round(success_rate, 2),
                'last_execution': stat.last_execution.strftime('%Y-%m-%d %H:%M:%S') if stat.last_execution else None
            })
        
        return jsonify({
            'code': 0,
            'data': {
                'metrics': {
                    'avg_latency': round(avg_latency, 2),
                    'packet_loss_rate': round(packet_loss_rate, 2),
                    'avg_jitter': round(avg_jitter, 2),
                    'availability': round(availability, 2)
                },
                'latency_trend': trend_data,
                'location_analysis': location_analysis,
                'task_details': task_details
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f"Error in get_ping_report: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取Ping报表数据失败: {str(e)}'
        }), 500


@bp.route('/reports/http', methods=['GET'])
@token_required
def get_http_report():
    """获取HTTP专项报表数据"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        task_id = request.args.get('task_id', type=int)
        status_code = request.args.get('status_code', type=str)
        
        # 计算时间范围
        end_time = datetime.now()
        if time_range == '1h':
            start_time = end_time - timedelta(hours=1)
        elif time_range == '1d':
            start_time = end_time - timedelta(days=1)
        elif time_range == '7d':
            start_time = end_time - timedelta(days=7)
        elif time_range == '30d':
            start_time = end_time - timedelta(days=30)
        else:
            start_time = end_time - timedelta(days=1)
        
        # 构建基础查询
        base_query = Result.query.join(Task, Result.task_id == Task.id).filter(
            Task.type == 'http',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        # HTTP关键指标统计
        total_requests = base_query.count()
        successful_requests = base_query.filter(Result.status == 'success').count()
        success_rate = (successful_requests / total_requests * 100) if total_requests > 0 else 0
        
        # 计算平均响应时间和吞吐量
        avg_response_time = 0
        avg_throughput = 0
        
        if total_requests > 0:
            response_times = []
            throughputs = []
            
            results = base_query.all()
            for result in results:
                if result.result_data:
                    try:
                        data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                        if 'response_time' in data:
                            response_times.append(data['response_time'])
                        if 'throughput' in data:
                            throughputs.append(data['throughput'])
                    except:
                        pass
            
            if response_times:
                avg_response_time = sum(response_times) / len(response_times)
            if throughputs:
                avg_throughput = sum(throughputs) / len(throughputs)
        
        error_rate = 100 - success_rate
        
        # HTTP响应时间趋势
        trend_data = []
        current_time = start_time
        interval = timedelta(hours=1) if time_range in ['1h', '1d'] else timedelta(days=1)
        
        while current_time < end_time:
            next_time = current_time + interval
            
            period_results = base_query.filter(
                Result.created_at >= current_time,
                Result.created_at < next_time
            ).all()
            
            period_response_times = []
            period_success = 0
            
            for result in period_results:
                if result.status == 'success':
                    period_success += 1
                if result.result_data:
                    try:
                        data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                        if 'response_time' in data:
                            period_response_times.append(data['response_time'])
                    except:
                        pass
            
            period_avg_response_time = sum(period_response_times) / len(period_response_times) if period_response_times else 0
            period_success_rate = (period_success / len(period_results) * 100) if period_results else 0
            
            trend_data.append({
                'time': current_time.strftime('%H:%M' if time_range in ['1h', '1d'] else '%m-%d'),
                'avg_response_time': round(period_avg_response_time, 2),
                'success_rate': round(period_success_rate, 2),
                'total': len(period_results)
            })
            
            current_time = next_time
        
        # 状态码分布统计
        status_code_stats = {}
        results = base_query.all()
        
        for result in results:
            if result.result_data:
                try:
                    data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                    if 'status_code' in data:
                        code = str(data['status_code'])
                        status_code_stats[code] = status_code_stats.get(code, 0) + 1
                except:
                    pass
        
        status_code_distribution = []
        for code, count in status_code_stats.items():
            percentage = (count / total_requests * 100) if total_requests > 0 else 0
            status_code_distribution.append({
                'status_code': code,
                'count': count,
                'percentage': round(percentage, 2)
            })
        
        # HTTP任务详细统计
        task_stats = db.session.query(
            Task.id,
            Task.name,
            Task.target,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success'),
            func.max(Result.created_at).label('last_execution')
        ).join(Result, Task.id == Result.task_id).filter(
            Task.type == 'http',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            task_stats = task_stats.filter(Task.id == task_id)
        
        task_stats = task_stats.group_by(Task.id, Task.name, Task.target).all()
        
        task_details = []
        for stat in task_stats:
            success_rate = (stat.success / stat.total * 100) if stat.total > 0 else 0
            task_details.append({
                'task_id': stat.id,
                'task_name': stat.name,
                'target': stat.target,
                'total': stat.total,
                'success': stat.success,
                'success_rate': round(success_rate, 2),
                'last_execution': stat.last_execution.strftime('%Y-%m-%d %H:%M:%S') if stat.last_execution else None
            })
        
        return jsonify({
            'code': 0,
            'data': {
                'metrics': {
                    'success_rate': round(success_rate, 2),
                    'avg_response_time': round(avg_response_time, 2),
                    'avg_throughput': round(avg_throughput, 2),
                    'error_rate': round(error_rate, 2)
                },
                'response_time_trend': trend_data,
                'status_code_distribution': status_code_distribution,
                'task_details': task_details
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f"Error in get_http_report: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取HTTP报表数据失败: {str(e)}'
        }), 500


@bp.route('/reports/api', methods=['GET'])
@token_required
def get_api_report():
    """获取API专项报表数据"""
    try:
        # 获取查询参数
        time_range = request.args.get('time_range', '1d', type=str)
        task_id = request.args.get('task_id', type=int)
        transaction_type = request.args.get('transaction_type', type=str)
        
        # 计算时间范围
        end_time = datetime.now()
        if time_range == '1h':
            start_time = end_time - timedelta(hours=1)
        elif time_range == '1d':
            start_time = end_time - timedelta(days=1)
        elif time_range == '7d':
            start_time = end_time - timedelta(days=7)
        elif time_range == '30d':
            start_time = end_time - timedelta(days=30)
        else:
            start_time = end_time - timedelta(days=1)
        
        # 构建基础查询
        base_query = Result.query.join(Task, Result.task_id == Task.id).filter(
            Task.type == 'api',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            base_query = base_query.filter(Task.id == task_id)
        
        # API关键指标统计
        total_transactions = base_query.count()
        successful_transactions = base_query.filter(Result.status == 'success').count()
        transaction_success_rate = (successful_transactions / total_transactions * 100) if total_transactions > 0 else 0
        
        # 计算断言通过率和平均事务时间
        assertion_pass_rate = 0
        avg_transaction_time = 0
        
        if total_transactions > 0:
            assertion_passes = 0
            transaction_times = []
            
            results = base_query.all()
            for result in results:
                if result.result_data:
                    try:
                        data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                        if 'assertion_passed' in data and data['assertion_passed']:
                            assertion_passes += 1
                        if 'transaction_time' in data:
                            transaction_times.append(data['transaction_time'])
                    except:
                        pass
            
            assertion_pass_rate = (assertion_passes / total_transactions * 100) if total_transactions > 0 else 0
            if transaction_times:
                avg_transaction_time = sum(transaction_times) / len(transaction_times)
        
        failure_rate = 100 - transaction_success_rate
        
        # API事务成功率趋势
        trend_data = []
        current_time = start_time
        interval = timedelta(hours=1) if time_range in ['1h', '1d'] else timedelta(days=1)
        
        while current_time < end_time:
            next_time = current_time + interval
            
            period_results = base_query.filter(
                Result.created_at >= current_time,
                Result.created_at < next_time
            ).all()
            
            period_success = 0
            period_assertion_passes = 0
            
            for result in period_results:
                if result.status == 'success':
                    period_success += 1
                if result.result_data:
                    try:
                        data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                        if 'assertion_passed' in data and data['assertion_passed']:
                            period_assertion_passes += 1
                    except:
                        pass
            
            period_success_rate = (period_success / len(period_results) * 100) if period_results else 0
            period_assertion_rate = (period_assertion_passes / len(period_results) * 100) if period_results else 0
            
            trend_data.append({
                'time': current_time.strftime('%H:%M' if time_range in ['1h', '1d'] else '%m-%d'),
                'transaction_success_rate': round(period_success_rate, 2),
                'assertion_pass_rate': round(period_assertion_rate, 2),
                'total': len(period_results)
            })
            
            current_time = next_time
        
        # 失败原因分析
        failure_reasons = {
            '断言失败': 0,
            '接口超时': 0,
            '服务器错误': 0,
            '参数错误': 0,
            '网络异常': 0,
            '其他错误': 0
        }
        
        failed_results = base_query.filter(Result.status != 'success').all()
        for result in failed_results:
            if result.result_data:
                try:
                    data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                    if 'error_type' in data:
                        error_type = data['error_type']
                        if error_type in failure_reasons:
                            failure_reasons[error_type] += 1
                        else:
                            failure_reasons['其他错误'] += 1
                    else:
                        failure_reasons['其他错误'] += 1
                except:
                    failure_reasons['其他错误'] += 1
        
        failure_analysis = []
        total_failures = sum(failure_reasons.values())
        for reason, count in failure_reasons.items():
            percentage = (count / total_failures * 100) if total_failures > 0 else 0
            failure_analysis.append({
                'reason': reason,
                'count': count,
                'percentage': round(percentage, 2)
            })
        
        # API任务详细统计
        task_stats = db.session.query(
            Task.id,
            Task.name,
            Task.target,
            func.count(Result.id).label('total'),
            func.sum(func.case([(Result.status == 'success', 1)], else_=0)).label('success'),
            func.max(Result.created_at).label('last_execution')
        ).join(Result, Task.id == Result.task_id).filter(
            Task.type == 'api',
            Result.created_at >= start_time,
            Result.created_at <= end_time
        )
        
        if task_id:
            task_stats = task_stats.filter(Task.id == task_id)
        
        task_stats = task_stats.group_by(Task.id, Task.name, Task.target).all()
        
        task_details = []
        for stat in task_stats:
            success_rate = (stat.success / stat.total * 100) if stat.total > 0 else 0
            
            # 获取该任务的性能分析数据
            task_results = base_query.filter(Task.id == stat.id).all()
            
            # 计算性能指标
            response_times = []
            assertion_passes = 0
            step_breakdown_data = []
            
            for result in task_results:
                if result.result_data:
                    try:
                        data = json.loads(result.result_data) if isinstance(result.result_data, str) else result.result_data
                        
                        # 收集响应时间
                        if 'response_time' in data:
                            response_times.append(data['response_time'])
                        elif 'transaction_time' in data:
                            response_times.append(data['transaction_time'])
                        
                        # 统计断言通过
                        if 'assertion_passed' in data and data['assertion_passed']:
                            assertion_passes += 1
                        
                        # 收集步骤分解数据
                        if 'step_breakdown' in data and isinstance(data['step_breakdown'], list):
                            step_breakdown_data.extend(data['step_breakdown'])
                    except:
                        pass
            
            # 计算性能统计
            avg_response_time = sum(response_times) / len(response_times) if response_times else 0
            response_times_sorted = sorted(response_times) if response_times else [0]
            p95_index = int(len(response_times_sorted) * 0.95) if response_times_sorted else 0
            p99_index = int(len(response_times_sorted) * 0.99) if response_times_sorted else 0
            p95_response_time = response_times_sorted[p95_index] if p95_index < len(response_times_sorted) else avg_response_time * 1.5
            p99_response_time = response_times_sorted[p99_index] if p99_index < len(response_times_sorted) else avg_response_time * 2
            
            assertion_pass_rate = (assertion_passes / stat.total * 100) if stat.total > 0 else 0
            
            # 处理步骤分解数据
            step_summary = []
            if step_breakdown_data:
                step_dict = {}
                for step in step_breakdown_data:
                    if isinstance(step, dict) and 'name' in step and 'time' in step:
                        step_name = step['name']
                        step_time = step['time']
                        if step_name not in step_dict:
                            step_dict[step_name] = []
                        step_dict[step_name].append(step_time)
                
                for step_name, times in step_dict.items():
                    avg_time = sum(times) / len(times) if times else 0
                    step_summary.append({
                        'name': step_name,
                        'avg_time': round(avg_time, 2)
                    })
            
            task_details.append({
                'task_id': stat.id,
                'task_name': stat.name,
                'target': stat.target,
                'total': stat.total,
                'success': stat.success,
                'success_rate': round(success_rate, 2),
                'assertion_pass_rate': round(assertion_pass_rate, 2),
                'avg_response_time': round(avg_response_time, 2),
                'p95_response_time': round(p95_response_time, 2),
                'p99_response_time': round(p99_response_time, 2),
                'step_breakdown': step_summary,
                'last_execution': stat.last_execution.strftime('%Y-%m-%d %H:%M:%S') if stat.last_execution else None
            })
        
        return jsonify({
            'code': 0,
            'data': {
                'metrics': {
                    'transaction_success_rate': round(transaction_success_rate, 2),
                    'assertion_pass_rate': round(assertion_pass_rate, 2),
                    'avg_transaction_time': round(avg_transaction_time, 2),
                    'failure_rate': round(failure_rate, 2)
                },
                'success_rate_trend': trend_data,
                'task_list': task_details,
                'failure_reasons': failure_analysis,
                'transaction_success_trend': trend_data,
                'failure_analysis': failure_analysis,
                'task_details': task_details
            },
            'message': 'success'
        })
        
    except Exception as e:
        print(f"Error in get_api_report: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'获取API报表数据失败: {str(e)}'
        }), 500


@bp.route('/reports/<int:report_id>/export', methods=['POST'])
@token_required
def export_report(report_id):
    """导出报表"""
    try:
        data = request.get_json()
        export_format = data.get('format', 'excel')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        filters = data.get('filters', {})
        
        # 根据report_id确定报表类型
        report_type_map = {
            1: 'overview',
            2: 'tcp', 
            3: 'ping',
            4: 'http',
            5: 'api'
        }
        report_type = report_type_map.get(report_id, 'overview')
        
        # 生成Excel文件
        if export_format == 'excel':
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f'{report_type.upper()}报表'
            
            # 设置标题
            title = f'{report_type.upper()}专项报表'
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:F1')
            
            # 设置时间范围
            time_range = f'时间范围: {start_date} 至 {end_date}'
            ws['A2'] = time_range
            ws['A2'].font = Font(size=12)
            ws.merge_cells('A2:F2')
            
            # 设置表头
            headers = ['时间', '总数', '成功数', '失败数', '成功率(%)', '平均响应时间(ms)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # 添加示例数据（实际应用中应该从数据库查询）
            sample_data = [
                ['2025-08-27 00:00', 100, 95, 5, 95.0, 120.5],
                ['2025-08-27 01:00', 98, 92, 6, 93.9, 125.3],
                ['2025-08-27 02:00', 102, 98, 4, 96.1, 118.7],
                ['2025-08-27 03:00', 95, 90, 5, 94.7, 122.1],
                ['2025-08-27 04:00', 105, 100, 5, 95.2, 119.8]
            ]
            
            for row_idx, row_data in enumerate(sample_data, 5):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 调整列宽
            column_letters = ['A', 'B', 'C', 'D', 'E', 'F']
            for i, letter in enumerate(column_letters):
                max_length = 0
                for row in ws.iter_rows(min_col=i+1, max_col=i+1):
                    for cell in row:
                        if hasattr(cell, 'value') and cell.value is not None:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                adjusted_width = min(max(max_length + 2, 12), 25)
                ws.column_dimensions[letter].width = adjusted_width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 生成文件名
            filename = f'{report_type}_report_{start_date}_{end_date}.xlsx'
            
            # 返回文件
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        else:
            # 暂不支持其他格式
            return jsonify({
                'code': 1,
                'data': None,
                'message': '暂不支持该导出格式'
            }), 400
        
    except Exception as e:
        print(f"Error in export_report: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'code': 1,
            'data': None,
            'message': f'导出报表失败: {str(e)}'
        }), 500